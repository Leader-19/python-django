import os
import tempfile
import subprocess
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def convert_excel_to_pdf(file):
    with tempfile.TemporaryDirectory() as tmp:
        input_path = os.path.join(tmp, file.name)
        
        # Save uploaded file
        with open(input_path, "wb+") as f:
            for chunk in file.chunks():
                f.write(chunk)

        pdf_filename = os.path.splitext(file.name)[0] + ".pdf"
        output_pdf = os.path.join(tmp, pdf_filename)

        try:
            # === 1. Pre-process with openpyxl (most important part) ===
            wb = load_workbook(input_path)
            
            for ws in wb.worksheets:
                # Landscape + Paper size
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                
                # FORCE Fit to Width (this is critical)
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                
                # Very small margins to give maximum space
                ws.page_margins.left = 0.1
                ws.page_margins.right = 0.1
                ws.page_margins.top = 0.3
                ws.page_margins.bottom = 0.3
                ws.page_margins.header = 0.1
                ws.page_margins.footer = 0.1

                # Auto-adjust column widths (prevent too wide columns)
                for col in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col)
                    # Make columns reasonable before scaling
                    current_width = ws.column_dimensions[col_letter].width or 12
                    ws.column_dimensions[col_letter].width = min(18, current_width)

            wb.save(input_path)

            # === 2. Convert with aggressive PDF export options ===
            subprocess.run([
                "soffice",
                "--headless",
                "--nologo",
                "--nolockcheck",
                "--convert-to",
                'pdf:calc_pdf_Export:{"PageLayout":{"type":"long","value":1},"UseLosslessCompression":{"type":"boolean","value":"true"}}',
                "--outdir", tmp,
                input_path
            ], check=True, capture_output=True, text=True)

            if os.path.exists(output_pdf):
                with open(output_pdf, "rb") as f:
                    response = HttpResponse(f.read(), content_type="application/pdf")
                    response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
                    return response
            else:
                return HttpResponse("PDF not generated", status=500)

        except subprocess.CalledProcessError as e:
            error = e.stderr or e.stdout or str(e)
            return HttpResponse(f"LibreOffice Error: {error}", status=500)
        except Exception as e:
            return HttpResponse(f"Error: {str(e)}", status=500)